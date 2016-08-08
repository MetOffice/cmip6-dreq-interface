# Extracting miptable maps from the DREQ using djq and the spreadsheet
# for comparisons
#

__all__ = ('DJQMap', 'XLSXMap',
           'empty_miptables', 'prune_empty_miptables',
           'miptables_with_empty_mipsets', 'prune_empty_mipsets')

from os.path import join
from collections import defaultdict
from re import split
from djq import (process_request, default_dqroot, default_dqtag, ensure_dq,
                 valid_dqtag, valid_dqroot)
import djq.variables.cv_invert_varmip as civ

from openpyxl import load_workbook

class MTMap(object):
    # superclass for miptable maps
    def __init__(self, tag=None, root=None,
                 *args, **kwargs):
        # Is there any part of Python's argument defaulting and object
        # system which is not broken? How did it get to be this way?
        assert valid_dqtag(tag), "bad tag"
        assert valid_dqroot(root), "bad root"
        self.tag = tag if tag is not None else default_dqtag()
        self.root = root if root is not None else default_dqroot()

# Computing what djq thinks is going on
#

class DJQMap(MTMap):
    # This is just wrapping up various functions (as static methods)
    # and caching the result
    def __init__(self, cvimpl=civ, *args, **kwargs):
        super(DJQMap, self).__init__(*args, **kwargs)
        self.dq = ensure_dq(dqtag=self.tag, dqroot=self.root)
        self.cvimpl = cvimpl
        self.__results = None

    @staticmethod
    def jsonify(dq, cmvids):
        # a minimal JSONifier (much faster than the default one)
        def jc(cmvid):
            cmv = dq.inx.uid[cmvid]
            return {'label': cmv.label,
                    'uid': cmvid,
                    'miptable': cmv.mipTable}
        return tuple(jc(cmvid) for cmvid in cmvids)

    @staticmethod
    def invert_replies(replies):
        table = defaultdict(lambda: defaultdict(set))
        for reply in replies:
            for vt in reply['reply-variables']:
                table[vt['miptable']][vt['uid']].add(reply['mip'])
        # returning defaultdicts can be confusing
        return dict(table)

    @property
    def results(self):
        # mindlessly cached results
        if self.__results is not None:
            return self.__results
        else:
            dq = self.dq
            self.__results = self.invert_replies(
                process_request(tuple({'mip': mip.label,
                                       'experiment': True}
                                      for mip in dq.coll['mip'].items),
                                dq=dq,
                                cvimpl=self.cvimpl,
                                jsimpl=self.jsonify))
            return self.__results

# Now do the same from the spreadsheet
#

class XLSXMap(MTMap):
    def __init__(self, skip = {'Notes'}, *args, **kwargs):
        super(XLSXMap, self).__init__(*args, **kwargs)
        self.wb = load_workbook(
            filename=join(self.root, "tags", self.tag,
                          "dreqPy/docs/CMIP6_MIP_tables.xlsx"),
            read_only=True)
        # this is just for inspectability: it is not used
        self.headers = tuple(
            (i, c.value)
            for (i, c) in enumerate(tuple(self.wb['Amon'].rows)[0]))
        assert len(self.headers) == 24, "wrong number of columns"
        self.skip = skip
        self.__results = None

    @staticmethod
    def setify_mips(mips):
        # turn a stringy list of MIPs from the spreadsheet into a
        # set
        return set(split(r",\s*", mips)
                   if mips is not None
                   else ())

    @property
    def results(self):
        # The columns we want are 18 (UID), 22 and 23 (the two lists
        # of MIPs), and again we want to build a map from {miptable:
        # {uid: mips}}
        if self.__results is not None:
            return self.__results
        else:
            wb = self.wb
            setify = self.setify_mips
            skip = self.skip
            self.__results = {miptable: {r[18].value: (setify(r[22].value)
                                                       | setify(r[23].value))
                                         for r in tuple(wb[miptable].rows)[1:]}
                              for miptable in wb.sheetnames
                              if miptable not in skip}
            return self.__results

# a miptable is empty if it has no variables, or if all its variables
# have empty sets of MIPs.

def empty_miptables(mtmap):
    return set((miptable
                for (miptable, variables) in mtmap.iteritems()
                if (len(variables) == 0
                    or max(map(len, variables.values())) == 0)))

def prune_empty_miptables(mtmap):
    return {miptable: variables
            for (miptable, variables) in mtmap.iteritems()
            if (len(variables) > 0
                and max(map(len, variables.values())) > 0)}


# A miptable has empty mipsets if any of its variables do
#

def has(pred, iterable):
    for i in iterable:
        if pred(i):
            return True
    return False

def miptables_with_empty_mipsets(mtmap):
    return set((miptable
                for (miptable, variables) in mtmap.iteritems()
                if has(lambda(s): len(s) == 0, variables.values())))

def prune_empty_mipsets(mtmap):
    empties = miptables_with_empty_mipsets(mtmap)
    def prune_maybe(mt, variables):
        if mt in empties:
            return {uid: mipset for (uid, mipset) in variables.iteritems()
                    if len(mipset) > 0}
        else:
            return variables
    return {miptable: prune_maybe(miptable, variables)
            for (miptable, variables) in mtmap.iteritems()}
