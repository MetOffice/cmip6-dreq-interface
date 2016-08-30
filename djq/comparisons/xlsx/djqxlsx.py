# (C) British Crown Copyright 2016, Met Office.
# See LICENSE.md in the top directory for license details.
#

# Extracting miptable maps from the DREQ using djq and the spreadsheet
# for comparisons
#
# This is experimental code, and horrid in various places
#

__all__ = ('DJQMap', 'XLSXMap')

from os.path import join
from collections import defaultdict
from re import split
from djq import (process_request, default_dqroot, default_dqtag, ensure_dq,
                 valid_dqtag, valid_dqroot)
import djq.variables.cv_invert_varmip as civ

from openpyxl import load_workbook

def has(pred, iterable):
    # does any element of an iterable matche a predicate?
    for i in iterable:
        if pred(i):
            return True
    return False

class MTMap(object):
    # superclass for miptable maps
    #
    # The interesting attributes are:
    # - root & tag are what you think
    # - name is a name for the object
    # - results is a dict mapping miptables to dicts mapping CMORvar
    #   UIDs to sets of MIPs
    #
    def __init__(self, tag=None, root=None, name=None,
                 *args, **kwargs):
        # Is there any part of Python's argument defaulting and object
        # system which is not broken? How did it get to be this way?
        super(MTMap, self).__init__(*args, **kwargs)
        self.root = root if root is not None else default_dqroot()
        self.tag = tag if tag is not None else default_dqtag()
        assert valid_dqroot(self.root), "bad root"
        assert valid_dqtag(self.tag), "bad tag"
        self.name = name
        self.__results = None

    # Caching results.  This is pretty bogus
    #
    @property
    def results(self):
        if self.__results is None:
            self.__results = self.compute_results()
        return self.__results

    @results.setter
    def results(self, value):
        self.__results = value

    @property
    def dq(self):
        return ensure_dq(dqtag=self.tag, dqroot=self.root)

    def missing_uids(self):
        # Return the UIDs missing from the dreq in self's results
        inx = self.dq.inx.uid
        return set(uid
                   for vs in self.results.itervalues()
                   for uid in vs.iterkeys()
                   if uid not in inx)

    def prune_missing_uids(self):
        inx = self.dq.inx.uid
        return {miptable: {variable: mipset
                           for (variable, mipset) in variables.iteritems()
                           if variable in inx}
                for (miptable, variables) in self.results.iteritems()}

    def empty_miptables(self):
        # empty miptables of a map
        return set((miptable
                    for (miptable, variables) in self.results.iteritems()
                    if (len(variables) == 0
                        or max(map(len, variables.values())) == 0)))

    def prune_empty_miptables(self):
        return {miptable: variables
                for (miptable, variables) in self.results.iteritems()
                if (len(variables) > 0
                    and max(map(len, variables.values())) > 0)}


    def miptables_with_empty_mipsets(self):
        # miptables containing variables with empty mipsets
        return set((miptable
                    for (miptable, variables) in self.results.iteritems()
                    if has(lambda(s): len(s) == 0, variables.values())))

    def prune_empty_mipsets(self):
        empties = self.miptables_with_empty_mipsets()
        def prune_maybe(mt, variables):
            if mt in empties:
                return {uid: mipset for (uid, mipset) in variables.iteritems()
                        if len(mipset) > 0}
            else:
                return variables
        return {miptable: prune_maybe(miptable, variables)
                for (miptable, variables) in self.results.iteritems()}

# Computing what djq thinks is going on
#

class DJQMap(MTMap):
    # This is just wrapping up various functions (as static methods)
    # and caching the result
    def __init__(self, cvimpl=civ, name="djq", *args, **kwargs):
        super(DJQMap, self).__init__(*args, name=name, **kwargs)
        self.cvimpl = cvimpl
        self.__results = None

    @staticmethod
    def jsonify(dq, cmvids):
        # a minimal JSONifier (much faster than the default one)
        def jc(cmvid):
            cmv = dq.inx.uid[cmvid]
            return {'label': cmv.label,
                    'uid': cmvid,
                    'miptable': cmv.mipTable}
        return tuple(jc(cmvid) for cmvid in cmvids)

    @staticmethod
    def invert_replies(replies):
        table = defaultdict(lambda: defaultdict(set))
        for reply in replies:
            for vt in reply['reply-variables']:
                table[vt['miptable']][vt['uid']].add(reply['mip'])
        # returning defaultdicts can be confusing (note we have to
        # convert two levels)
        return {mt: dict(vs) for (mt, vs) in table.iteritems()}

    def compute_results(self):
        # compute results (cached by MTMap)
        dq = self.dq
        return self.invert_replies(
            process_request(tuple({'mip': mip.label, 'experiment': True}
                                  for mip in dq.coll['mip'].items),
                            dq=dq,
                            cvimpl=self.cvimpl,
                            jsimpl=self.jsonify))

# Now do the same from the spreadsheet
#

class XLSXMap(MTMap):
    def __init__(self, skip={'Notes'}, name="xlsx", *args, **kwargs):
        super(XLSXMap, self).__init__(*args, name=name, **kwargs)
        self.wb = load_workbook(
            filename=join(self.root, "tags", self.tag,
                          "dreqPy/docs/CMIP6_MIP_tables.xlsx"),
            read_only=True)
        # this is just for inspectability: it is not used
        self.headers = tuple(
            (i, c.value)
            for (i, c) in enumerate(tuple(self.wb['Amon'].rows)[0]))
        assert len(self.headers) == 24, "wrong number of columns"
        self.skip = skip
        self.__results = None

    @staticmethod
    def setify_mips(mips):
        # turn a stringy list of MIPs from the spreadsheet into a
        # set
        return set(split(r",\s*", mips)
                   if mips is not None
                   else ())

    def compute_results(self):
        # The columns we want are 18 (UID), 22 and 23 (the two lists
        # of MIPs), and again we want to build a map from {miptable:
        # {uid: mips}}
        wb = self.wb
        setify = self.setify_mips
        skip = self.skip
        return {miptable: {r[18].value: (setify(r[22].value)
                                         | setify(r[23].value))
                           for r in tuple(wb[miptable].rows)[1:]}
                for miptable in wb.sheetnames
                if miptable not in skip}
